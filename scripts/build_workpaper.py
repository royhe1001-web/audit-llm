#!/usr/bin/env python3
"""
阶段四·步骤 9: Excel 工作底稿
============================
8 个 Sheet:
  1. 数据概览
  2. 财务特征分析 (描述统计 + 相关性)
  3. 模型评估
  4. Top100 高风险
  5. 规则触发清单
  6. 行业分析
  7. 异常检测
  8. 仪表盘说明
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

BASE = "/Users/Zhuanz/claude工作文件夹/审计数据分析大作业"
DATA_DIR = os.path.join(BASE, "data")
OUT_DIR = os.path.join(BASE, "output")

FIN_COLS = ['roe', 'roa', 'debt_ratio', 'current_ratio', 'asset_turnover', 'net_margin', 'ocf_to_rev']

# ============================================================
# 样式
# ============================================================
HEADER_FILL = PatternFill(start_color='1F3A5F', end_color='1F3A5F', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F3A5F')
CENTER = Alignment(horizontal='center', vertical='center')
WRAP = Alignment(wrap_text=True, vertical='top')


def write_sheet(wb, name, df, freeze='B2', col_widths=None):
    ws = wb.create_sheet(name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # 表头样式
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    # 冻结
    ws.freeze_panes = freeze
    # 列宽
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    else:
        for col in ws.columns:
            try:
                max_len = max(len(str(c.value)) for c in col if c.value)
                ws.column_dimensions[col[0].column_letter].width = min(max_len * 1.3, 30)
            except Exception:
                pass
    return ws


# ============================================================
# 1. 加载数据
# ============================================================
print("=" * 60)
print("步骤 9.1: 加载数据")
print("=" * 60)

enriched = pd.read_csv(os.path.join(DATA_DIR, "fraud_features_enriched.csv"))
risk = pd.read_csv(os.path.join(DATA_DIR, "risk_scored_full.csv"))
rules = pd.read_csv(os.path.join(OUT_DIR, "rule_trigger_aggregate.csv"))
anom = pd.read_csv(os.path.join(DATA_DIR, "anomaly_companies.csv"))
top100 = pd.read_csv(os.path.join(OUT_DIR, "top100_high_risk.csv"))
ind_xl = pd.read_excel(os.path.join(OUT_DIR, "category_layer_stats.xlsx"), sheet_name='行业')

print(f"  enriched: {enriched.shape}")
print(f"  risk: {risk.shape}")
print(f"  rules: {rules.shape}")
print(f"  anomalies: {anom.shape}")
print(f"  top100: {top100.shape}")
print(f"  industry: {ind_xl.shape}")

# ============================================================
# 2. 创建 Workbook
# ============================================================
wb = Workbook()
wb.remove(wb.active)

# ============================================================
# Sheet 1: 数据概览
# ============================================================
print("\n" + "=" * 60)
print("Sheet 1: 数据概览")
print("=" * 60)

# 关键指标
overview = pd.DataFrame({
    '指标': [
        '总行数', '唯一公司数', '唯一行业数', '唯一地区数',
        'ann_related=1 (年报相关违规)', 'ann_fin_flag=1 (财务信息影响)',
        'ann_fin_flag=0 (非财务信息影响)', 'ann_fin_flag=缺失 (未标注)',
        'third_party_flag=1 (涉及第三方)',
        '财务完整行数 (7特征全不缺失)', '财务完整率',
        '行业覆盖率 (匹配 tushare)',
        '年份范围', '平均风险分数', '高风险数', '中风险数', '低风险数', '异常数',
    ],
    '值': [
        len(enriched), enriched['Symbol'].nunique(), enriched['industry'].nunique(), enriched['area'].nunique(),
        (enriched['ann_related'] == 1).sum(),
        (enriched['ann_fin_flag'] == 1).sum(),
        (enriched['ann_fin_flag'] == 0).sum(),
        enriched['ann_fin_flag'].isna().sum(),
        (enriched['third_party_flag'] == 1).sum(),
        enriched['roe'].notna().sum(),
        f"{enriched['roe'].notna().mean()*100:.1f}%",
        f"{(enriched['industry'] != '未分类').mean()*100:.1f}%",
        f"{int(enriched['violation_year'].min())}-{int(enriched['violation_year'].max())}",
        f"{risk['risk_score'].mean():.3f}",
        (risk['risk_level'] == '高风险').sum(),
        (risk['risk_level'] == '中风险').sum(),
        (risk['risk_level'] == '低风险').sum(),
        anom.shape[0],
    ],
})
write_sheet(wb, '1_数据概览', overview, freeze='A2',
            col_widths={'A': 35, 'B': 30})

# ============================================================
# Sheet 2: 财务特征分析
# ============================================================
print("\nSheet 2: 财务特征分析")

# 描述统计
desc = enriched[FIN_COLS].describe().T.round(4)
desc = desc.reset_index().rename(columns={'index': '财务特征'})
write_sheet(wb, '2_财务_描述统计', desc, freeze='B2',
            col_widths={'A': 18})

# 相关性
corr = enriched[FIN_COLS].corr().round(3)
corr = corr.reset_index().rename(columns={'index': '财务特征'})
write_sheet(wb, '2_财务_相关性', corr, freeze='B2')

# 各特征 top 10 极值
for col in FIN_COLS[:3]:  # 演示 3 个
    s = enriched[['ShortName', 'Symbol', 'industry', 'violation_year', col]].dropna()
    s = s.sort_values(col, ascending=False).head(20).reset_index(drop=True)
    s.columns = ['公司名', '代码', '行业', '违规年份', col]
    write_sheet(wb, f'2_财务_{col}_TOP20', s, freeze='A2')

# ============================================================
# Sheet 3: 模型评估
# ============================================================
print("\nSheet 3: 模型评估")

metrics = pd.DataFrame({
    '指标': ['Accuracy', 'F1 Score', 'Recall', 'Precision', 'AUC', '5-Fold CV F1 (mean)', '5-Fold CV F1 (std)'],
    '测试集值': [0.810, 0.762, 0.747, 0.780, 0.840, 0.762, 0.025],
    '说明': [
        '整体准确率', '正例 F1 分数(主要指标)',
        '正例召回率(查全率)', '正例精确率(查准率)',
        'ROC 曲线下面积', '5 折交叉验证 F1 平均', '5 折交叉验证 F1 标准差',
    ],
})
write_sheet(wb, '3_模型评估', metrics, freeze='A2',
            col_widths={'A': 25, 'B': 15, 'C': 35})

# 特征重要性
import joblib
pipe = joblib.load(os.path.join(BASE, "models/risk_scoring_pipeline.pkl"))
rf = pipe['model'].named_steps['clf']
imp_df = pd.DataFrame({
    '特征': pipe['fin_cols'],
    '重要性': rf.feature_importances_.round(4),
}).sort_values('重要性', ascending=False).reset_index(drop=True)
write_sheet(wb, '3_特征重要性', imp_df, freeze='A2')

# ============================================================
# Sheet 4: Top100 高风险
# ============================================================
print("\nSheet 4: Top100 高风险")

display_cols = ['ShortName', 'Symbol', 'industry', 'violation_year',
                'p_ml', 'n_rules_triggered', 'rule_ids', 'risk_score', 'risk_level', 'ann_fin_flag']
write_sheet(wb, '4_Top100高风险', top100[display_cols], freeze='B2')

# ============================================================
# Sheet 5: 规则触发清单
# ============================================================
print("\nSheet 5: 规则触发清单")

write_sheet(wb, '5_规则触发_汇总', rules, freeze='C2')

# 规则严重度说明
rule_doc = pd.DataFrame({
    '规则ID': ['R1', 'R2', 'R3', 'R4', 'R5', 'R6'],
    '规则名': ['consecutive_loss', 'cashflow_divergence', 'high_leverage',
               'liquidity_stress', 'asset_turnover_extreme', 'roe_anomaly'],
    '中文名': ['连续亏损', '现金流背离', '高负债', '流动性紧张', '资产周转异常', 'ROE 异常'],
    '严重度': [25, 30, 15, 20, 10, 25],
    '触发条件': [
        'roe<0 AND net_margin<0',
        'ocf_to_rev<0 AND net_margin>0',
        'debt_ratio>0.7',
        'current_ratio<1',
        'asset_turnover<0.3 OR asset_turnover>5',
        'roa>0.1 AND roe<0',
    ],
    '审计逻辑': [
        '经营性造血能力缺失,舞弊动机高',
        '利润高但无现金回流,盈余操纵高发',
        '偿债压力大,易触发违规融资',
        '短期偿债能力不足,破产/重组预警',
        '经营效率极端,需下钻',
        '杠杆异常或少数股东掏空',
    ],
})
write_sheet(wb, '5_规则说明', rule_doc, freeze='A2',
            col_widths={'A': 10, 'B': 22, 'C': 12, 'D': 8, 'E': 35, 'F': 40})

# ============================================================
# Sheet 6: 行业分析
# ============================================================
print("\nSheet 6: 行业分析")

write_sheet(wb, '6_行业风险排名', ind_xl, freeze='B2')

# 行业风险高的 Top10
top_ind = ind_xl.sort_values('mean_risk', ascending=False).head(15)
top_ind['排名'] = range(1, len(top_ind) + 1)
write_sheet(wb, '6_行业风险_TOP15', top_ind[['排名'] + list(top_ind.columns[:-1])], freeze='A2')

# ============================================================
# Sheet 7: 异常检测
# ============================================================
print("\nSheet 7: 异常检测")

anom_display = anom[['ShortName', 'Symbol', 'industry', 'violation_year',
                       'risk_score', 'risk_level', 'p_ml', 'rule_ids',
                       'anomaly_method', 'ann_fin_flag']].copy()
anom_display = anom_display.sort_values('risk_score', ascending=False).reset_index(drop=True)
anom_display['排名'] = anom_display.index + 1
write_sheet(wb, '7_异常公司', anom_display, freeze='B2')

# 异常方法统计
anom_summary = pd.DataFrame({
    '方法': ['IF+LOF 双标记', '仅 IF', '仅 LOF', '合计'],
    '数量': [
        (anom['anomaly_method'] == 'IF+LOF').sum(),
        (anom['anomaly_method'] == 'IF_only').sum(),
        (anom['anomaly_method'] == 'LOF_only').sum(),
        len(anom),
    ],
    '占比': [
        f"{(anom['anomaly_method']=='IF+LOF').mean()*100:.1f}%",
        f"{(anom['anomaly_method']=='IF_only').mean()*100:.1f}%",
        f"{(anom['anomaly_method']=='LOF_only').mean()*100:.1f}%",
        '100.0%',
    ],
})
write_sheet(wb, '7_异常方法统计', anom_summary, freeze='A2')

# ============================================================
# Sheet 8: 仪表盘说明
# ============================================================
print("\nSheet 8: 仪表盘说明")

dash_doc = pd.DataFrame({
    'Tab': ['1. 总览', '2. 行业分析', '3. 公司画像', '4. Top100', '5. 异常检测'],
    '关键图表': [
        'KPI 卡 × 5 + 风险等级饼图 + 风险分数直方图 + ML vs 规则气泡图',
        '行业风险排名柱状图 + Top 5 行业时间趋势 + 行业-年份热力图',
        '7 维财务雷达图 + 6 规则红绿灯 + 历史违规记录',
        'Top100 表格(可筛选) + RF 特征重要性 + Top100 行业分布',
        'PCA 散点图 + 异常公司列表 + 异常方法多选筛选',
    ],
    '核心交互': [
        '侧边栏全局筛选(行业/年份/风险等级)',
        '5 行业下钻 + 12 年时间窗',
        '单选下拉选公司,自动渲染画像',
        '表格列可排序,行可点击',
        'IF+LOF/IF_only/LOF_only 多选',
    ],
    '启动命令': ['streamlit run app/dashboard.py'] * 5,
    '访问URL': ['http://localhost:8501'] * 5,
})
write_sheet(wb, '8_仪表盘说明', dash_doc, freeze='A2',
            col_widths={'A': 12, 'B': 50, 'C': 35, 'D': 35, 'E': 25})

# 系统信息
sys_info = pd.DataFrame({
    '项目': [
        '项目名称', '作者学号', '作者姓名', '完成日期',
        '技术栈', '模型', '交叉验证', '数据量', '标签', '产出物',
    ],
    '详情': [
        '上市公司财务舞弊智能识别 — 持续审计系统',
        '2023111180',
        '何其轩',
        '2026-06-06',
        'Python 3.13 + scikit-learn 1.8 + streamlit 1.58 + plotly 6.7 + tushare',
        'Random Forest 300树, max_depth=8, class_weight=balanced',
        '5-Fold CV F1 = 0.762 ± 0.025',
        '8,724 标注 / 7,997 特征 / 2,842 公司',
        'ann_fin_flag (是否影响财务信息)',
        '中间表 + 仪表盘 + 工作底稿 + 综合报告',
    ],
})
write_sheet(wb, '8_系统信息', sys_info, freeze='A2',
            col_widths={'A': 15, 'B': 60})

# ============================================================
# 保存
# ============================================================
print("\n" + "=" * 60)
print("保存 Excel")
print("=" * 60)

out_path = os.path.join(OUT_DIR, "audit_workpaper.xlsx")
wb.save(out_path)
print(f"  → {out_path}")
print(f"  文件大小: {os.path.getsize(out_path)/1024:.0f} KB")
print(f"  Sheet 数: {len(wb.sheetnames)}")
print(f"  Sheet 列表: {wb.sheetnames}")

print("\n" + "=" * 60)
print("✅ 步骤 9 完成")
print("=" * 60)
